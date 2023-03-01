from datetime import datetime
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from docx2pdf import convert
from pypdf import PdfMerger
import zipfile
import os


# Instantiates variables for testing
def main():
    roster_and_grades = 'BMRA Roster and Grades - 11023.0001.xlsx'
    save_directory = 'Certs'
    doc = 'Certificate of Training - Edit.docx'
    create_docs(roster_and_grades, save_directory, doc,)

# Creates the docs w/correct names and converts them to PDF
def create_docs(roster_and_grades, save_directory, doc, code_override=None):

    # Load in worksheet and document
    doc = DocxTemplate(doc)
    wb = load_workbook(roster_and_grades)
    ws = wb.active

    # Keep track of course code, cert num, and the name for the combined file
    course_code_save = str(ws["A2"].value)
    combined_pdf_name = ""
    certificate_number = 1

    # Loop through rows in ws
    for row in ws['2:35']:

        # Data scraped from the excel
        course_code = str(row[0].value)
        first_name = row[2].value
        last_name = row[4].value
        # grade = row[5].value
        course_name = row[6].value
        start_date = (row[7].value)
        end_date = (row[8].value)
        # agency = row[9].value
        location = row[10].value
        # class_hours = row[11].value
        clps = row[12].value

        # Formating datetime
        if (isinstance(start_date, datetime)):
            start_date = start_date.strftime('%#m/%#d/%Y')
            end_date = end_date.strftime('%#m/%#d/%Y')

        # Creates a dictionary from the excel and allows for rendering of template
        template_fill = {}
        if(type(first_name) == str):
            if (end_date == start_date):
                template_fill = {
                'FNAME': first_name,
                'LNAME': last_name,
                'COURSE': course_name,
                'CLPS': clps,
                'LOCATION': location,
                'START_DATE': start_date,
            }
            else:
                end_date = '- ' + end_date
                template_fill = {
                    'FNAME': first_name,
                    'LNAME': last_name,
                    'COURSE': course_name,
                    'CLPS': clps,
                    'LOCATION': location,
                    'START_DATE': start_date,
                    'END_DATE': end_date
                }

            # Creates the docs in desired directory
            doc.render(template_fill)
            if (code_override != None):
                document_name = save_directory + "/" + str(certificate_number).zfill(2) + " - " + first_name + " " + last_name + " - Certificate of Training - " + code_override + ".docx"
            else:
                document_name = save_directory + "/" + str(certificate_number).zfill(2) + " - " + first_name + " " + last_name + " - Certificate of Training - " + course_code + ".docx"
            doc.save(document_name)
            certificate_number += 1

    # conversion of directory to pdf       
    convert(save_directory)
    docxs = get_docx(save_directory)
    for document in docxs:
        os.remove(document)
    wb.close()

     # Naming for the combined file
    if (code_override != None):
        combined_pdf_name = "Certificates of Training - " + code_override 
    else:
        combined_pdf_name = "Certificates of Training - " + course_code_save

    # Get pdf iterable
    pdfs = get_pdf(save_directory)
    # Create the output filename for compressed file
    output_filename = save_directory + "/Compressed " + combined_pdf_name + ".zip"
    # Instantiate zipfile
    zf = zipfile.ZipFile(output_filename, "w")

    # Make combined pdf and compressed file
    merger = PdfMerger()
    for pdf in pdfs:
        zf.write(pdf)
        merger.append(pdf)
    
    # Make file
    merger.write(save_directory + "/" + combined_pdf_name + ".pdf")
    merger.close()
    zf.close()
    

# creates an interable of the docx files in folder
def get_docx(save_directory):
    return (os.path.join(save_directory, file)
    for file in os.listdir(save_directory)
    if 'docx' in file)

# creates an interable of the pdf files in folder
def get_pdf(save_directory):
    return (os.path.join(save_directory, file)
    for file in os.listdir(save_directory)
    if 'pdf' in file)

if __name__ == "__main__":
    main()
